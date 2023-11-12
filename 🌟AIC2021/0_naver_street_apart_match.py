# -*- coding: utf-8 -*-

import pickle

with open("도로명_리스트.pickle", "rb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    street = pickle.load(pickle_data)
    
# import random
# street = random.sample(street, 20)
# with open("USEDSTREET.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
#     pickle.dump(street, pickle_data)


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
import re
import pyautogui

browser = webdriver.Chrome("/Users/kunwoosmac/Downloads/chromedriver")

def Apart_Name(street_name):
    try:
        # 네이버에 검색해서 해당 아파트 이름 가져오기
        browser.get("https://www.naver.com")
        try:
            WebDriverWait(browser,10).until(ec.presence_of_element_located(By.ID, "query"))
        except:
            time.sleep(1)

        elem = browser.find_element_by_id("query")
        elem.send_keys(street_name)
        browser.find_element_by_id("search_btn").click()

        try:
            WebDriverWait(browser,10).until(ec.presence_of_element_located(By.CLASS_NAME, "ITiBH"))
        except:
            time.sleep(1)
        # print(browser.find_element_by_class_name("ITiBH").text)
        # print(browser.find_elements_by_class_name("_3rnws")[0].text)
        # print(browser.find_elements_by_class_name("_3rnws")[1].text)
        add = browser.find_element_by_class_name("ITiBH").text
        add = re.split(' ', add)[1]
        apart_name = browser.find_elements_by_class_name("_3rnws")[1].text
        apart_name = re.split('지번', apart_name)[0]
        add2 = browser.find_elements_by_class_name("_3rnws")[1].text
        add2 = re.split('서울특별시', add2)[1]
        match_add = "서울시" + str(re.split('\d', add2)[0])
        apart_name = str(add + apart_name)

        return apart_name, match_add
    except:
        return "Error", "Error"

def Search_Real(apart_name, match_add):
    try:
        # 부동산에서 검색
        browser.get("https://land.naver.com/")
        time.sleep(1)
        elem = browser.find_element_by_id("queryInputHeader")
        elem.send_keys(apart_name)
        browser.find_element_by_xpath("/html/body/div[4]/div[2]/div[1]/div/div[1]/div/fieldset/a[1]").click()

        time.sleep(1)

        # 단지정보 클릭
        try:
            browser.find_element_by_xpath("/html/body/div[2]/div/section/div[2]/div[1]/div/div[2]/div[1]/div[1]/div[2]/div[2]/button[1]").click()
            time.sleep(1)
        except:
            add = []
            elem = browser.find_elements_by_class_name("address")
            for i in range(0, len(elem)):
                add.append(re.sub(' ','', elem[i].text))
            # print(add)

            matchadd = re.sub(' ','',match_add)

            for i in range(0, len(add)):
                if add[i] == matchadd:
                    # print(add[i], match_add)
                    # print("YES")
                    try:
                        browser.find_elements_by_class_name("address")[i].click()
                    except:
                        pass

            time.sleep(1)
            try:
                browser.find_element_by_xpath(\
                    "/html/body/div[2]/div/section/div[2]/div[1]/div/div[2]/div[1]/div[1]/div[2]/div[2]/button[1]").click()
            except:
                pyautogui.moveTo(x=133, y=452)
                pyautogui.click()

            time.sleep(1)

        #########################################
        info = browser.find_element_by_id("detailContents1").text
        #info = browser.find_element_by_class_name("info_table_wrap").text
        #####요기서 제대로 못불러옴!!!!

        info = re.split('\n', info)
        info2 = []

        for i in range(0, len(info)):
            info2.extend(re.split(' ', info[i]))
        print(info2)

        saedae = info2[info2.index("세대수")+1]
        parking = info2[info2.index("총주차대수") + 1] + info2[info2.index("총주차대수") + 2]
        yong = info2[info2.index("용적률") + 1]
        gun = info2[info2.index("건폐율") + 1]
        print(saedae, parking, yong, gun)
        return saedae, parking, yong, gun
    
    except Exception as err:
        print(err)
        return "Error","Error","Error","Error"

i = 1
whole_data = {}
err_list = []

from openpyxl import load_workbook

apart_name, match_add = Apart_Name("4.19로11길 16")
print(Search_Real(apart_name, match_add))

# wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/2_도로별_데이터_추가.xlsx")
# ws = wb.active
# ws.cell(row = 1, column = 11, value = "세대수")
# ws.cell(row = 1, column = 12, value = "주차공간")
# ws.cell(row = 1, column = 13, value = "용적률")
# ws.cell(row = 1, column = 14, value = "건폐율")

# for street_name in street:
#     print(street_name, i)
#     i += 1
#     apart_name, match_add = Apart_Name(street_name)
#     saedae, parking, yong, gun = Search_Real(apart_name, match_add)
#     if saedae == "Error":
#         err_list.append(street_name)
#     else:
#         nar_data = {"세대수" : saedae, "주차공간" : parking, "용적률" : yong, "건폐율" : gun}
#         whole_data[apart_name] = nar_data
#         ws.cell(row = i, column = 11, value = saedae)
#         ws.cell(row = i, column = 12, value = parking)
#         ws.cell(row = i, column = 13, value = yong)
#         ws.cell(row = i, column = 14, value = gun)

# print(len(err_list))

# with open("아파트별_데이터_사전.pickle", "wb") as pickle_data: 
#     pickle.dump(whole_data, pickle_data)

# with open("아파트_데이터_에러.pickle", "wb") as pickle_data: 
#     pickle.dump(err_list, pickle_data)

# wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/3_도로별_데이터_추가.xlsx")
# wb.close()