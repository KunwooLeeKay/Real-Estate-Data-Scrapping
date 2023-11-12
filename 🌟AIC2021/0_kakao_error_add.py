import requests
from urllib.parse import urlparse
import pandas as pd
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
import re

# with open("에러난도로명리스트.pickle", "rb") as pickle_data:
#     err = pickle.load(pickle_data)


def search_Station(lon, lat):
    df = pd.DataFrame(columns = ['place_name','road_address_name', 'distance', 'x', 'y'])
    page = 1
    while True:
        url = "https://dapi.kakao.com/v2/local/search/category.json?&category_group_code=SW8&x="\
        +str(lon)+"&y="+str(lat)+"&page="+str(page)+"&radius=500" # 반경 500m
        json_obj = requests.get(urlparse(url).geturl(),headers={"Authorization":"KakaoAK 4f9b52bce91e7316a53be6acaeeed3b4"}).json()
        for document in json_obj['documents']:
            df_s = pd.DataFrame(document, index=[0])[['place_name','road_address_name', 'distance', 'x', 'y']]
            df = df.append(df_s)
        if json_obj['meta']['is_end'] == False:
            page += 1
        else:
            return df


whole_data = {}
count = 0

wb = load_workbook("학군_지하철_실거래가.xlsx")
ws = wb.active
streets = []
for i in range(2, ws.max_row + 1):
    streets.append(ws.cell(row = i, column = 1).value)

print(streets)


for i in range(0, len(streets)):
    print(i)
    
    address = streets[i]
    if ws.cell(row = i+2, column = 2).value != "N/A" and ws.cell(row=i+2, column = 6).value:
        try:
            keyword = ws.cell(row = i+2, column = 6).value
            if "현" in keyword:
                keyword = re.split("현", keyword)[1]
            print(keyword)
            page=1 # 첫번째 페이지
            url = "https://dapi.kakao.com//v2/local/search/keyword.json?query=" + str(keyword)
            json_obj = requests.get(urlparse(url).geturl(),headers={"Authorization":"KakaoAK 4f9b52bce91e7316a53be6acaeeed3b4"}).json()
            x = json_obj['documents'][0]['x']
            y = json_obj['documents'][0]['y']
            print(x,y)

            df = search_Station(x, y)

            ws.cell(row = i + 2, column = 5, value = len(df))
        except:
            whole_data[address] = "Error"
            ws.cell(row = i + 2, column = 5, value = "Error")
            count += 1


wb.save("Final_학군_지하철_실거래가.xlsx")
wb.close()

print(count)