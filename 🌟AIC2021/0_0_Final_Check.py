from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/3_도로별_데이터_추가.xlsx")
ws = wb.active

data3 = {}

for i in range(2, ws.max_row + 1):
    temp = []
    for j in range(11, 15):
        temp.append(ws.cell(row = i, column = j).value)
    data3[ws.cell(row = i, column = 1).value] = temp

print(data3["강남대로162길 27-26"])
wb.close()

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/5_도로별_데이터_추가.xlsx")
ws = wb.active

data5 = {}
for i in range(2, ws.max_row + 1):
    temp = []
    for j in range(12, 16):
        temp.append(ws.cell(row = i, column = j).value)
    data5[ws.cell(row = i, column = 1).value] = temp

print(data5["가마산로 276"])

street = list(data5.keys())


for i in range(0, len(data5)):
    if data5[street[i]][1] == "0대(세대당0.0대)":
        if data3[street[i]][1] != None:
            data5[street[i]][1] = data3[street[i]][1]
        else:
            data5[street[i]][1] = '-'

    if data5[street[i]][2] == "0%":
        if data3[street[i]][2] != None:
            data5[street[i]][2] = data3[street[i]][2]
            data5[street[i]][3] = data3[street[i]][3]
        else:
            data5[street[i]][2] = '-'
            data5[street[i]][3] = '-'
    
    if data5[street[i]][0] == "Value_Error":
        for j in range(0,4):
            if data3[street[i]][j] != None:
                data5[street[i]][j] = data3[street[i]][j]
            else:
                data5[street[i]][j] = '-'

for i in range(2, ws.max_row + 1):
    ws.cell(row = i, column = 12, value = data5[ws.cell(row = i, column = 1).value][0])
    ws.cell(row = i, column = 13, value = data5[ws.cell(row = i, column = 1).value][1])
    ws.cell(row = i, column = 14, value = data5[ws.cell(row = i, column = 1).value][2])
    ws.cell(row = i, column = 15, value = data5[ws.cell(row = i, column = 1).value][3])

# for i in range(2, ws.max_row + 1):
#     if ws.cell(row = i, column = 2).value == "N/A":
#         ws.delete_rows(i)


wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/6_도로별_데이터_추가.xlsx")
wb.close()
