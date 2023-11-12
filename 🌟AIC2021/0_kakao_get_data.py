import requests
from urllib.parse import urlparse
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
import re
# with open("에러난도로명리스트.pickle", "rb") as pickle_data:
#     err = pickle.load(pickle_data)


def search_Station(lon, lat):
    df = []
    page = 1
    while True:
        url = "https://dapi.kakao.com/v2/local/search/category.json?&category_group_code=SW8&x="\
        +str(lon)+"&y="+str(lat)+"&page="+str(page)+"&radius=500" # 반경 500m
        json_obj = requests.get(urlparse(url).geturl(),headers={"Authorization":"KakaoAK 4f9b52bce91e7316a53be6acaeeed3b4"}).json()
        i=0
        for document in json_obj['documents']:
            df.append(json_obj['documents'][i]['place_name'])
            i = i + 1
        if json_obj['meta']['is_end'] == False:
            page += 1
        else:
            return df


whole_data = {}
count = 0

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/1_도로별_데이터_추가(학교).xlsx")
ws = wb.active
streets = []
for i in range(2, ws.max_row + 1):
    streets.append(ws.cell(row = i, column = 1).value)

print(streets)

for i in range(0, len(streets)):
    print(i)
    address = streets[i]

    try:
        page=1 # 첫번째 페이지
        url = "https://dapi.kakao.com//v2/local/search/address.json?query=" + str(address)
        json_obj = requests.get(urlparse(url).geturl(),headers={"Authorization":"KakaoAK 4f9b52bce91e7316a53be6acaeeed3b4"}).json()

        x = json_obj['documents'][0]['x']
        y = json_obj['documents'][0]['y']

        df = search_Station(x, y)
        whole_data[address] = df

        subways = ""
        # ws.cell(row = i + 2, column = 5, value = len(df))
        for name in df:
            subways = subways + name
            subways = subways +','
        ws.cell(row = i + 2, column = 10, value = subways)
    except:
        if ws.cell(row = i+2, column = 2).value != "N/A" and ws.cell(row=i+2, column = 6).value:
            try:
                keyword = ws.cell(row = i+2, column = 6).value
                if "현" in keyword:
                    keyword = re.split("현", keyword)[1]
                print(keyword)
                page=1 # 첫번째 페이지
                url = "https://dapi.kakao.com//v2/local/search/keyword.json?query=" + str(keyword)
                json_obj = requests.get(urlparse(url).geturl(),headers={"Authorization":"KakaoAK 4f9b52bce91e7316a53be6acaeeed3b4"}).json()
                x = json_obj['documents'][0]['x']
                y = json_obj['documents'][0]['y']

                df = search_Station(x, y)

                # ws.cell(row = i + 2, column = 5, value = len(df))
                subways = ""
                for name in df:
                    subways = subways + name
                    subways = subways +','
                ws.cell(row = i + 2, column = 10, value = subways)
            except:
                whole_data[address] = ["Error"]
                ws.cell(row = i + 2, column = 10, value = "Error")
                count += 1
        elif ws.cell(row = i+2, column = 2).value != "N/A":
            pass
        else:
            whole_data[address] = ["Error"]
            ws.cell(row = i + 2, column = 10, value = "Error")
            count += 1


wb.save("2_도로별_데이터_추가.xlsx")
wb.close()

with open("Subway_Data.pickle", "wb") as pickle_data:
    pickle.dump(whole_data, pickle_data)

print(count) # should be 198