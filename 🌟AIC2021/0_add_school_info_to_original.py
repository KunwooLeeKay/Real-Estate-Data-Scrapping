from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb = load_workbook("/Users/kunwoosmac/Downloads/도로별_학군추가_에러수정본.xlsx")
ws = wb.active

mid = {}
high = {}

for i in range(2, ws.max_row + 1):
    mid[ws.cell(row = i, column = 1).value] = ws.cell(row = i, column = 3).value
    high[ws.cell(row = i, column = 1).value] = ws.cell(row = i, column = 4).value

wb.close
# print(mid)
# print(high)

for year in range(2013, 2021):
    wb = Workbook()
    wb = load_workbook("아파트(전월세)_실거래가_"+ str(year) + ".xlsx")
    ws = wb.active
    ws.cell(row = 17, column = 16, value = "중학교 수")
    ws.cell(row = 17, column = 17, value = "고등학교 수")
    for i in range(18, ws.max_row + 1):
        street = ws.cell(row = i, column = 14).value
        m = mid.get(street)
        h = high.get(street)

        if m == None or h == None:
            m = "학군 정보에 없음"
            h = "학군 정보에 없음"

        ws.cell(row = i, column = 16, value = m)
        ws.cell(row = i, column = 17, value = h)
        print(str(street) + ' : ' + str(m)+','+str(h))
    wb.save(str(year) + "_실거래가_학군정보추가.xlsx")
    wb.close


