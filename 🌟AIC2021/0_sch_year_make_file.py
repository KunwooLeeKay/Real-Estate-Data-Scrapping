from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/도로별_데이터_추가.xlsx")
ws = wb.active
wb.close()

import pickle
with open("도로별_중학교_이름.pickle", 'rb') as pickle_data:
    mid = pickle.load(pickle_data)

with open("도로별_고등학교_이름.pickle", 'rb') as pickle_data:
    high = pickle.load(pickle_data)

with open("도로별_고등학교_이름.pickle", 'rb') as pickle_data:
    high = pickle.load(pickle_data)

with open("중학교_개교.pickle", "rb") as pickle_data: 
    mid_year = pickle.load(pickle_data)

with open("고등학교_개교.pickle", "rb") as pickle_data: 
    high_year = pickle.load(pickle_data)

# 잘못된 정보 수정

print("Error_can't_search" in mid_year.values())
print("Error_different_form" in mid_year.values())
lst = list(mid_year.values())
for i in range(0, len(lst)):
    if lst[i] == "Error_different_form":
        print(i) # 8, 103, 369

school = list(mid_year.keys())
for i in [8, 103, 369]:
    print(school[i]) # 강빛중학교, 마곡하늬중학교, 항동중학교

mid_year["강빛중학교"] = "2021년 3월 1일"
mid_year["마곡하늬중학교"] = "2020년 3월 1일"
mid_year["항동중학교"] = "2020년 3월 1일"

print("Error_can't_search" in mid_year.values())
print("Error_different_form" in mid_year.values())

print("Error_can't_search" in high_year.values())
print("Error_different_form" in high_year.values())


# 학교 이름이랑 년도 따로 할당
midschool = list(mid_year.keys())
middate = list(mid_year.values())
midyear = []
import re
for i in range(0, len(middate)):
    midyear.append(re.split('년', middate[i])[0])

highschool = list(high_year.keys())
highdate = list(high_year.values())
highyear = []

# 2012이후 설립학교 담을 사전
sub_mid = {}
sub_high = {}

for i in range(0, len(highdate)):
    highyear.append(re.split('년', highdate[i])[0])
# print(midyear, highyear)
for i in range(0, len(midyear)):
    if int(midyear[i]) > 2012:
        print(midschool[i])
        sub_mid[midschool[i]] = midyear[i]
for i in range(0, len(highyear)):
    if int(highyear[i]) > 2012:
        print(highschool[i])
        sub_high[highschool[i]] = highyear[i]

print(sub_mid)
print(sub_high)


street = []
add_mid = ""
add_high = ""
for i in range(2, ws.max_row + 1):
    print(i)
    street.append(ws.cell(row = i, column = 1).value)
    if ws.cell(row = i, column = 2).value != "N/A" and ws.cell(row = i, column = 3).value != "N/A":
        mid_list = mid[ws.cell(row = i, column = 1).value]
        if type(mid_list) == list:
            year = []
            year_string = ''
            for j in range(0, len(mid_list)):
                if mid_list[j] in list(sub_mid.keys()):
                    year_string = year_string + sub_mid[mid_list[j]] + '/'
            ws.cell(row = i, column = 8, value = year_string)
        else:
            ws.cell(row = i, column = 8, value = "Error")
        
        high_list = high[ws.cell(row = i, column = 1).value]
        if type(high_list) == list:
            year = []
            year_string = ''
            for j in range(0, len(high_list)):
                if high_list[j] in list(sub_high.keys()):
                    year_string = year_string + sub_high[high_list[j]] + '/'
            ws.cell(row = i, column = 9, value = year_string)
        else:
            ws.cell(row = i, column = 9, value = "Error")

wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/1_도로별_데이터_추가(학교).xlsx")
wb.close()
