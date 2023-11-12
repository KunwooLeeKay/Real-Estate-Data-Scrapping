from selenium import webdriver
import time
import re
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook

def Get_Data(apartment_name):
    try:
        browser.get("https://realty.daum.net/home/apt")
        elem = browser.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div\
            /div[2]/div[1]/div[1]/div/div[1]/input")
        elem.send_keys(apartment_name)
        time.sleep(2)

        elem.send_keys(Keys.ENTER)
        elem.send_keys(Keys.ENTER)

        time.sleep(1)

        browser.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[2]/div/div/div[1]/div[2]/\
             div[2]/div/div[3]/div[1]/div[1]/div/div[1]/div/div[1]/div/div[1]").click()

        parkinglot = "Can't Click More Err"
    except:
        return "검색 안됨"

    # 전체 정보를 가져오는것으로 변경..!
    try:
        parkinglot = browser.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[2]/div/div/div[1]/div[2]/div[2]/div/div[3]/div[1]/div[2]/div").text
        return parkinglot
    except:
        return "정보 못가져옴"

browser = webdriver.Chrome("/Users/kunwoosmac/Downloads/chromedriver")
browser.get("https://realty.daum.net/home/apt/")
browser.maximize_window()
elem = browser.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div[1]/div/div[1]/input")


wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성

wb = load_workbook("Renew_Apartment_list.xlsx")
ws = wb.active
whole_data = {}
##
for i in range(2, ws.max_row + 1, 1000):
    print(i)
    apartment_name = ws.cell(row = i, column = 1).value
    data = Get_Data(apartment_name)

    if data == "검색 안됨":
        change_apartment_name = re.sub("\(.*\)|\s-\s.*",'',apartment_name)
        data = Get_Data(change_apartment_name)

    elif data == "정보 못가져옴":
        pass

    sep_data = re.split("\n", data)
    nar_data = {}


    for name in ['세대 수', '시공사', '동수', '주차', '난방', '편의시설', '용적률/건폐율']:
        if name in sep_data:
            nar_data[name] = sep_data[sep_data.index(name)+1]
        elif name not in sep_data:
            nar_data[name] = "N/A"

    whole_data[apartment_name] = nar_data

import pickle
with open("아파트별_데이터_사전.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    pickle.dump(whole_data, pickle_data)

##
for i in range(2, ws.max_row + 1,1000):
    ws.cell(row = i, column = 2, value = whole_data[ws.cell(row = i, column = 1).value]['세대 수'])
    ws.cell(row = i, column = 3, value = whole_data[ws.cell(row = i, column = 1).value]['시공사'])
    ws.cell(row = i, column = 4, value = whole_data[ws.cell(row = i, column = 1).value]['동수'])
    ws.cell(row = i, column = 5, value = whole_data[ws.cell(row = i, column = 1).value]['주차'])
    ws.cell(row = i, column = 6, value = whole_data[ws.cell(row = i, column = 1).value]['난방'])
    if whole_data[ws.cell(row = i, column = 1).value]['편의시설'] == "N/A":
        conv = "N/A"
    else:
        conv = len(re.split(',',whole_data[ws.cell(row = i, column = 1).value]['편의시설']))
    ws.cell(row = i, column = 7, value = conv)
    if whole_data[ws.cell(row = i, column = 1).value]['용적률/건폐율'] != "N/A":
        new = re.split('/', whole_data[ws.cell(row = i, column = 1).value]['용적률/건폐율'])
        yong = re.sub(' ','',new[0])
        gun = re.sub(' ','',new[1])
    else:
        yong = gun = "N/A"
    ws.cell(row = i, column = 8, value = yong)
    ws.cell(row = i, column = 9, value = gun)


wb.save("아파트별_건물정보추가_실거래가.xlsx")
wb.close()


