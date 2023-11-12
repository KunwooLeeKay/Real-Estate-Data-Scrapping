from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb = load_workbook("Renew_Street_Name.xlsx")

ws = wb.active

street = []

for i in range (209, ws.max_row + 1):
    street.append(ws.cell(row = i, column = 1).value)

wb.close
wb = Workbook()

old_street = []
wb = load_workbook("Street_Name.xlsx")
ws = wb.active
for i in range (2, ws.max_row + 1):
    old_street.append(ws.cell(row=i, column = 1).value)

wb.close

pop_list = []

for i in range(0, len(street)):
    if street[i] in old_street:
        pop_list.append(i)
print(len(pop_list))
print(len(street))
print(len(old_street))
pop_list.sort()
pop_list.reverse()
print(pop_list)
for i in range(0, len(pop_list)):
    street.pop(pop_list[i])

wb = Workbook()
wb = load_workbook("도로별_학군추가_실거래가.xlsx")
ws = wb.active

for i in range(1, len(street) + 1):
    ws.cell(row = i + 2802, column = 1, value = street[i-1])
wb.save("Missed_Street_Names.xlsx")
wb.close