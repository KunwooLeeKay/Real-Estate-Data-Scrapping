import pickle

with open("중학교_개교.pickle", "rb") as pickle_data:
    mid = pickle.load(pickle_data)

with open("고등학교_개교.pickle", "rb") as pickle_data:
    high = pickle.load(pickle_data)

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

mid["강빛중학교"] = "2021년 3월 1일"
mid["마곡하늬중학교"] = "2020년 3월 1일"
mid["항동중학교"] = "2020년 3월 1일"

o_mid_school = list(mid.keys())
o_mid_year = list(mid.values())
mid_school = []
mid_year = []
import re
for i in range(0, len(o_mid_school)):
    if int(re.split("년", o_mid_year[i])[0])>=2012:
        mid_school.append(o_mid_school[i])
        mid_year.append(o_mid_year[i])
        

o_high_school = list(high.keys())
o_high_year = list(high.values())
high_school = []
high_year = []
for i in range(0, len(o_high_school)):
    if int(re.split("년", o_high_year[i])[0])>=2012:
        high_school.append(o_high_school[i])
        high_year.append(o_high_year[i])

for i in range(1, len(mid_school)+1):
    ws.cell(row = i, column = 1, value = mid_school[i-1])
    ws.cell(row = i, column = 2, value = mid_year[i-1])

for i in range(1, len(high_school)+1):
    ws.cell(row = i, column = 4, value = high_school[i-1])
    ws.cell(row = i, column = 5, value = high_year[i-1])

wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/개교일.xlsx")
wb.close()