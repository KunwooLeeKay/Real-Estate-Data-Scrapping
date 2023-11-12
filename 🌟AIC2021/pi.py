from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/아파트(전월세)_실거래가_2012.xlsx")
ws = wb.active
j = 0 
for i in range(18, ws.max_row+1):
    if ws.cell(row = i, column = 14).value == None:
        j = j + 1

print(j)
wb.close()