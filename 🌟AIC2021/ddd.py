from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/6_도로별_데이터_추가.xlsx")
whole_data = {}
ws = wb.active

for i in range(2, ws.max_row + 1):
    data_list = []
    data_list.append(ws.cell(row = i, column = 3).value) # whole_data["street"][0] = mid
    data_list.append(ws.cell(row = i, column = 4).value) # whole_data["street"][1] = high
    data_list.append(ws.cell(row = i, column = 5).value) # whole_data["street"][2] = subway in 500m
    data_list.append(ws.cell(row = i, column = 8).value) # whole_data["street"][3] = 중학교 뺄 년도
    data_list.append(ws.cell(row = i, column = 9).value) # whole_data["street"][4] = 고등학교 뺄 년도
    data_list.append(ws.cell(row = i, column = 11).value) # whole_data["street"][5] = 지하철 뺄 년도
    data_list.append(ws.cell(row = i, column = 12).value) # whole_data["street"][6] = 세대수
    data_list.append(ws.cell(row = i, column = 13).value) # whole_data["street"][7] = 주차공간
    data_list.append(ws.cell(row = i, column = 14).value) # whole_data["street"][8] = 용적률
    data_list.append(ws.cell(row = i, column = 15).value) # whole_data["street"][9] = 건폐율
    whole_data[ws.cell(row = i, column = 1).value] = data_list

print(whole_data["언주로 103"])

wb.close()

num = 0

print(int(num))
import re
year = 2012
try:
    lst = list(re.split('/', None))
    lst = list(filter(None, lst))
    sub = 0
    for j in range(0, len(lst)):
        if year > int(lst[j]):
            sub = sub - 1
    print(1)
except:
    if whole_data["언주로 103"][1] != None:
        print(whole_data["언주로 103"][2])
        print(2)
    else:

        print(3)