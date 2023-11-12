from selenium import webdriver
import re
browser = webdriver.Chrome("/Users/kunwoosmac/Downloads/chromedriver")
browser.get("https://schoolzone.emac.kr/gis/gis.do")

def Get_Data(street):

    browser.get("https://schoolzone.emac.kr/gis/gis.do")
    # 지역을 서울로 지정
    try:
        browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select").click()
        browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select/option[2]").click()
    except: 
        print("에러 line 17")
        return 0,0,0
    # 도로명 주소 검색부
    elem = browser.find_element_by_id("searchText")
    from selenium.webdriver.common.keys import Keys
    # 여기에 도로명 주소 입력해줘야함
    try:
        elem.send_keys(street)
    except:
        print(street + "sendkeys 오류!")
        return 0,0,0
    if len(street)<=2:
        return 0,0,0
    try:
        elem.send_keys(Keys.ENTER)
    except:
        print(street + "엔터 오류!")
        return 0,0,0

    # 팝업창 닫음
    try:
        browser.find_element_by_xpath("/html/body/div[1]/div[10]/div[1]/p/input").click()
        browser.find_element_by_xpath("/html/body/div[1]/div[9]/div[1]/p/input").click()
    except: pass

    browser.switch_to.frame(browser.find_element_by_name('searchIframe'))  # iframe있으면 이거 해줘야함..!

    # WebDriverWait(browser,10).until(ec.presence_of_element_located((By.CLASS_NAME, "btn_school_search1")))
    try:
        try:
            browser.find_element_by_xpath("/html/body/div[5]/div/ul[2]/li[1]/input").click()
            #/html/body/div[5]/div/ul[2]/li[1]/input
        except:
            browser.find_element_by_xpath("/html/body/div[5]/div[1]/ul[2]/li[1]/input").click()
    ### 검색 안될때!!!!!
    except:
        browser.get("https://schoolzone.emac.kr/gis/gis.do")
        # 지역을 서울로 지정
        try:
            browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select").click()
            browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select/option[2]").click()
        except: 
            print("에러 line 17")
            return 0,0,0
        elem = browser.find_element_by_id("searchText")

        if '-' in street:
            st = re.split('-', street)[0]
            street = st
        elif '길' in street:
            st = re.split('길', street)[0] + '길'
            street = st
        elem.send_keys(street)
        elem.send_keys(Keys.ENTER)
        try:
            try:
                browser.find_element_by_xpath("/html/body/div[5]/div/ul[2]/li[1]/input").click()
                #/html/body/div[5]/div/ul[2]/li[1]/input
            except:
                browser.find_element_by_xpath("/html/body/div[5]/div[1]/ul[2]/li[1]/input").click()
        except:
            pass


    try:
        # iframe 리스트 : 
        # schoolAreaIframe, iframe_hakguInfo, iframe_schoolDetail2, iframe_schoolDetail3, iframe_schoolDetail4, iframe_schoolDetail5

        browser.switch_to.default_content() #iframe 나가기

        browser.switch_to.frame(browser.find_element_by_name('schoolAreaIframe')) # 학군 정보 있는 iframe

        # 초등학교
        try:
            elem = browser.find_element_by_xpath("/html/body/div[2]/div/dl/dd/table/tbody/tr/td[1]/a")
            school1 = [elem.text]
            browser.find_element_by_xpath("/html/body/div[2]/div/dl/dt/a").click()
            #print(school1)

            # 중학교
            browser.find_element_by_xpath("/html/body/div[3]/div/dl/dt/a").click()
            num = 1
            school2 = []
            while True:
                try:
                    elem = browser.find_element_by_xpath("/html/body/div[3]/div/dl/dd/table/tbody/tr["+str(num)+"]/td[1]/a")
                    school2.append(elem.text)
                    num += 1
                except:
                    break
            browser.find_element_by_xpath("/html/body/div[3]/div/dl/dt/a").click()
            school2 = list(filter(None, school2))
            #print(school2)

            # 고등학교
            browser.find_element_by_xpath("/html/body/div[4]/div/dl/dt/a").click()
            num = 2
            school3 = []
            while True:
                try:
                    elem = browser.find_element_by_xpath("/html/body/div[4]/div/dl/dd/table/tbody/tr["+str(num)+"]/td[1]/a")
                    school3.append(elem.text)
                    num += 1
                except:
                    break
            browser.find_element_by_xpath("/html/body/div[4]/div/dl/dt/a").click()
            school3 = list(filter(None, school3))
        except:
            print(street+"에러!")
            return 0,0,0
    except:
        return 0,0,0
        
    #print(school3)
    return school1, school2, school3


from openpyxl import load_workbook
from openpyxl import Workbook

wb = Workbook()
wb = load_workbook("Renew_Street_Name.xlsx")
ws = wb.active

street = []
for i in range(1, ws.max_row + 1):
    street.append(ws.cell(row = i, column = 1).value)

## 피클에 도로명 저장해놓기
import pickle
with open("도로명_리스트.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    pickle.dump(street, pickle_data)

school_list_elem = []
school_list_mid = []
school_list_high = []

for i in range(0, len(street)):
    print(i)
    elem, mid, high = Get_Data(street[i])
    try:
        ws.cell(row = i + 1, column = 2, value = len(elem))
        ws.cell(row = i + 1, column = 3, value = len(mid))
        ws.cell(row = i + 1, column = 4, value = len(high))
    except:
        ws.cell(row = i + 1, column = 2, value = "Err")
        ws.cell(row = i + 1, column = 3, value = "Err")
        ws.cell(row = i + 1, column = 4, value = "Err")

    try:
        school_list_elem.extend(elem)
        school_list_mid.extend(mid)
        school_list_high.extend(high)
    except:
        pass

    if elem == 0 or mid == 0 or high ==0:
        ws.cell(row = i + 2803, column = 2, value = "Err")
        ws.cell(row = i + 2803, column = 3, value = "Err")
        ws.cell(row = i + 2803, column = 4, value = "Err")

school_list_elem = set(school_list_elem)
school_list_elem = list(school_list_elem)
school_list_elem.sort()

school_list_mid = set(school_list_mid)
school_list_mid = list(school_list_mid)
school_list_mid.sort()

school_list_high = set(school_list_high)
school_list_high = list(school_list_high)
school_list_high.sort()

# print(school_list_elem)
# print(school_list_mid)
# print(school_list_high)

with open("초등학교_리스트.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    pickle.dump(school_list_elem, pickle_data)

with open("중학교_리스트.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    pickle.dump(school_list_mid, pickle_data)
    
with open("고등학교_리스트.pickle", "wb") as pickle_data: # profile.pickle 파일을 읽기로 읽어서 profile_with변수에 저장
    pickle.dump(school_list_high, pickle_data)


wb.save("Final_Final_도로별_학군추가_실거래가.xlsx")
wb.close

