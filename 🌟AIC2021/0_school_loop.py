# Q. 학교까지의 거리도 가져와야할까?
# 실거래가 정보 당시에 존재하던 학교인지 조회해야함. 학교 개교일 가져와서 그 전이면 삭제 필요

from selenium import webdriver

browser = webdriver.Chrome("/Users/kunwoosmac/Downloads/chromedriver")
browser.get("https://schoolzone.emac.kr/gis/gis.do")

def Get_Data(street):
    browser.get("https://schoolzone.emac.kr/gis/gis.do")
    # 지역을 서울로 지정
    try:
        browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select").click()
        browser.find_element_by_xpath("/html/body/div[1]/div[1]/div[1]/div/div/span[1]/select/option[2]").click()
    except: 
        print("에러 line 17")
        return 0,0,0
    # 도로명 주소 검색부
    elem = browser.find_element_by_id("searchText")
    from selenium.webdriver.common.keys import Keys
    # 여기에 도로명 주소 입력해줘야함
    try:
        elem.send_keys(street)
    except:
        print(street + "sendkeys 오류!")
        return 0,0,0
    if len(street)<=2:
        return 0,0,0
    try:
        elem.send_keys(Keys.ENTER)
    except:
        print(street + "엔터 오류!")
        return 0,0,0

    # 팝업창 닫음
    try:
        browser.find_element_by_xpath("/html/body/div[1]/div[10]/div[1]/p/input").click()
        browser.find_element_by_xpath("/html/body/div[1]/div[9]/div[1]/p/input").click()
    except: pass

    browser.switch_to.frame(browser.find_element_by_name('searchIframe'))  # iframe있으면 이거 해줘야함..!

    # WebDriverWait(browser,10).until(ec.presence_of_element_located((By.CLASS_NAME, "btn_school_search1")))
    try:
        search1 = browser.find_elements_by_class_name("btn_school_search1")
        search1[len(search1)-1].click() # 맨 마지막에 있는거 클릭해줌

        # iframe 리스트 : 
        # schoolAreaIframe, iframe_hakguInfo, iframe_schoolDetail2, iframe_schoolDetail3, iframe_schoolDetail4, iframe_schoolDetail5

        browser.switch_to.default_content() #iframe 나가기

        browser.switch_to.frame(browser.find_element_by_name('schoolAreaIframe')) # 학군 정보 있는 iframe

        # 초등학교
        try:
            elem = browser.find_element_by_xpath("/html/body/div[2]/div/dl/dd/table/tbody/tr/td[1]/a")
            school1 = [elem.text]
            browser.find_element_by_xpath("/html/body/div[2]/div/dl/dt/a").click()
            #print(school1)

            # 중학교
            browser.find_element_by_xpath("/html/body/div[3]/div/dl/dt/a").click()
            num = 1
            school2 = []
            while True:
                try:
                    elem = browser.find_element_by_xpath("/html/body/div[3]/div/dl/dd/table/tbody/tr["+str(num)+"]/td[1]/a")
                    school2.append(elem.text)
                    num += 1
                except:
                    break
            browser.find_element_by_xpath("/html/body/div[3]/div/dl/dt/a").click()
            school2 = list(filter(None, school2))
            #print(school2)

            # 고등학교
            browser.find_element_by_xpath("/html/body/div[4]/div/dl/dt/a").click()
            num = 2
            school3 = []
            while True:
                try:
                    elem = browser.find_element_by_xpath("/html/body/div[4]/div/dl/dd/table/tbody/tr["+str(num)+"]/td[1]/a")
                    school3.append(elem.text)
                    num += 1
                except:
                    break
            browser.find_element_by_xpath("/html/body/div[4]/div/dl/dt/a").click()
            school3 = list(filter(None, school3))
        except:
            print(street+"에러!")
            return 0,0,0
    except:
        return 0,0,0
        
    #print(school3)
    return school1, school2, school3

# 엑셀파일 불러오기
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성
street_name = "None"

wb = load_workbook("도로명_추출본.xlsx")
ws = wb.active
ws.cell(row = 1, column = 2, value= "초등학교 수")
ws.cell(row = 1, column = 3, value = "중학교 수")
ws.cell(row = 1, column = 4, value = "고등학교 수")
for x in range(2, ws.max_row + 1):
    print("총 : "+ str(ws.max_row) +"행")
    print("중" + str(x))
    street_name = ws.cell(row = x, column = 1).value
    elem_school, middle_school, high_school = Get_Data(street_name)
    if elem_school == 0 or middle_school == 0 or high_school ==0:
        ws.cell(row = x, column = 2, value = "Err")
        ws.cell(row = x, column = 3, value = "Err")
        ws.cell(row = x, column = 4, value = "Err")
    #print(elem_school,middle_school,high_school)
    else:
        ws.cell(row = x, column = 2, value = len(elem_school))
        ws.cell(row = x, column = 3, value = len(middle_school))
        ws.cell(row = x, column = 4, value = len(high_school))

        #일단 학교 이름을 각각 가져오는 부분은 빼고 갯수만 가져옴
        # for i in range(0,len(elem_school)):
        #     elem_school_excel = elem_school_excel+ ',' + elem_school[i]
        # for i in range(0, len(middle_school)):
        #     middle_school_excel = middle_school_excel + ',' + middle_school[i]
        # for i in range(0, len(high_school)):
        #     high_school_excel = high_school_excel + ',' + high_school[i]
        # ws.cell(row = x, column = 15) = elem_school_excel
        # ws.cell(row = x, column = 16) = middle_school_excel
        # ws.cell(row = x, column = 17) = high_school_excel
    wb.save("도로별_학군추가_실거래가.xlsx")
    wb.close()



