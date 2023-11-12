from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성함
year = 2012
apartment_list = []
for year in range(2012, 2022):
    print(year)
    wb = load_workbook("아파트(전월세)_실거래가_"+ str(year) + ".xlsx")
    ws = wb.active  
    for i in range(18, ws.max_row + 1):
        apart = ws.cell(row = i, column = 5).value
        # if "(" in apart:
        #     apart = re.split("(", apart)[0]
        # elif "동~" in apart:
        #     # apart = re.split(" ")
        #     pass
        apartment_list.append(apart)
        
apartment_list = set(apartment_list)
apartment_list = list(apartment_list)
apartment_list.sort()
print(apartment_list)
row_num = len(apartment_list)

wb.close()

wb = Workbook()
ws = wb.active # 현재 활성화된 sheet를 가져옴 -> ws에 저장
ws.title = "Apartment_List"

for x in range(1, row_num):
    ws.cell(column = 1, row = x, value = apartment_list[x-1])

wb.save("Renew_Apartment_list.xlsx")
wb.close()