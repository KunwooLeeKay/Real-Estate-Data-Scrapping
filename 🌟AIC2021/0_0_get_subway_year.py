# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import re

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/4_도로별_데이터_추가.xlsx")
ws = wb.active

lst = []

for i in range(2, ws.max_row+1):
    if ws.cell(row = i, column = 10).value:
        lst.extend(re.split(',', ws.cell(row = i, column = 10).value))

lst = list(filter(None, lst))
lst = set(lst)
lst = list(lst)
del lst[lst.index('Error')]
lst.sort()
# print(len(lst))
# print(lst)


wb.close()

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/지하철역_데이터.xlsx")
ws = wb.active

lst2 = []
txt = ''
dic = {}
for i in range(2, ws.max_row + 1):
    txt = ws.cell(row = i, column = 3).value + "역 "
    txt = txt + str(ws.cell(row = i, column = 2).value) + "호선"
    dic[txt] = ws.cell(row = i, column = 4).value
    lst2.append(txt)

lst2.sort()


wb.close()

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/지하철역_데이터_추가본.xlsx")
ws = wb.active

lst3 = []
for i in range(1, ws.max_row+1):
    if ws.cell(row = i, column = 2).value:
        dic[ws.cell(row = i, column = 1).value] = ws.cell(row = i, column = 2).value
    else:
        dic[ws.cell(row = i, column = 1).value] = "1000"

wb.close()

print(dic)
for i in range(0, len(lst)):
    if lst[i] in list(dic.keys()):
        pass
    else:
        print("Err!" + str(lst[i]))
whole_data = {}

err_list = []
for name in lst:
    if name in lst2: # 그 지하철 파일에 있으면 거기서 연도 가져옴
        whole_data[name] = dic.get(name)
    else: # 없으면 스크래핑하는거 짜야되니까 err_list에 담아둠
        err_list.append(name)


wb = Workbook()
ws = wb.active
x = 2
sub_sub = {}
for i in range(0, len(dic)):
    if int(list(dic.values())[i]) > 2012:
        ws.cell(row = x, column = 1, value = list(dic.keys())[i])
        ws.cell(row = x, column = 2, value = list(dic.values())[i])
        sub_sub[list(dic.keys())[i]] = list(dic.values())[i]
        x = x + 1

wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/빼야되는_지하철역_모음.xlsx")
wb.close()

print(sub_sub)

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/4_도로별_데이터_추가.xlsx")
ws = wb.active

for i in range(2, ws.max_row + 1):
    if ws.cell(row = i, column = 10).value != "Error":
        years = ''
        if ws.cell(row = i, column = 10).value:
            st_sub = re.split(',', ws.cell(row = i, column = 10).value)
            st_sub = list(filter(None, st_sub))
            for j in range(0, len(st_sub)):
                if st_sub[j] in list(sub_sub.keys()):
                    years = years + str(sub_sub[st_sub[j]]) + '/'
            ws.cell(row = i, column = 11, value = years)

wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/5_도로별_데이터_추가.xlsx")
wb.close()