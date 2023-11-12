from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성함

wb = load_workbook("Scrapping.xlsx")
ws = wb.active

apartment = []
regex = "\(.*\)|\s-\s.*"
for x in range(1, ws.max_row+1):
    apartment.append(re.sub(regex,'',ws.cell(row = x, column = 1).value))
apartment = list(filter(None,apartment))
apartment = set(apartment)
apartment = list(apartment)
apartment.sort()

print(apartment)
print(len(apartment))
wb.close()

wb = Workbook()
ws = wb.active # 현재 활성화된 sheet를 가져옴 -> ws에 저장
ws.title = "Apartment_List"

for x in range(1, len(apartment)):
    ws.cell(column = 1, row = x, value = apartment[x-1])

wb.save("아파트 이름 추출본.xlsx")
wb.close()