from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성함
year = 2012
street = []
for year in range(2012, 2022):
    print(year)
    wb = load_workbook("아파트(전월세)_실거래가_"+ str(year) + ".xlsx")
    ws = wb.active
    for i in range(18, ws.max_row + 1):
        street.append(ws.cell(row = i, column = 14).value)

street = set(street)
street = list(street)
street.sort()
print(len(street))

wb.close()

wb = Workbook()
ws = wb.active # 현재 활성화된 sheet를 가져옴 -> ws에 저장
ws.title = "Street_Name_List"

for x in range(0, len(street)):
    ws.cell(column = 1, row = x+1, value = street[x])

wb.save("Renew_Street_Name.xlsx")
wb.close()