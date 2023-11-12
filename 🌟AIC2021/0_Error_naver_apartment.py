from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import re


browser = webdriver.Chrome("/Users/kunwoosmac/Downloads/chromedriver")

street_list = []
row_list = []
wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/3_도로별_데이터_추가.xlsx")
ws = wb.active
for i in range(2, ws.max_row + 1):
    if ws.cell(row = i, column = 2).value != "N/A":
        if ws.cell(row = i, column = 11).value == None or ws.cell(row = i, column = 11).value == '-' or \
            ws.cell(row = i, column = 12).value == None or ws.cell(row = i, column = 12).value == '-' or \
                ws.cell(row = i, column = 13).value == None or ws.cell(row = i, column = 13).value == '-' or\
                    ws.cell(row = i, column = 14).value == None or ws.cell(row = i, column = 14).value == '-':
                    street_list.append(ws.cell(row = i, column = 1).value)
                    row_list.append(i)

print(len(street_list))

err_list = []

def Search_real(street_name):

    try:
        url = "https://land.seoul.go.kr:444/land/"
        browser.get(url)
        elem = browser.find_element_by_id("query")
        elem.send_keys(street_name)
        elem.send_keys(Keys.ENTER)
        time.sleep(2)
        i = 1
        z = 2
        while True:
            try:
                if i == 8:
                    browser.find_element_by_id("a" + str(z)).click()
                    time.sleep(2)
                    z = z + 1
                    i = 1

                elem = browser.find_element_by_xpath("/html/body/header/div[2]/div[3]/form/div/div[3]/ul["+str(i)+"]/li/div[1]/div[3]/span[2]")
                if street_name.replace(' ','') in elem.text.replace(' ',''):
                    browser.find_element_by_xpath("/html/body/header/div[2]/div[3]/form/div/div[3]/ul["+str(i)+"]/li/div[2]/a").click()
                    break
                i = i + 1
            except:
                break

        try:
            time.sleep(0.5)
            browser.find_element_by_id("tab2").click()
        except:
            time.sleep(1)
            browser.find_element_by_id("tab2").click()



        # 주건물 / 부건물 막 있으면 엉뚱한거 퍼올수도 있음
        time.sleep(1)
        i = 1
        while True: #elem.get_attribute("title") != "표제부(주건축물)":
            try:
                elem = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/section/table[1]/tbody/tr["+str(i)+"]/td[1]/div")

                if elem.get_attribute("title") =="총괄표제부" or elem.get_attribute("title") == "일반건축물(주건축물)":

                    elem.click()

                    break
                elif elem.get_attribute("title") == "표제부(주건축물)":
                    elem.click()
                    break

                i += 1
                if i == 10:
                    break
            except:

                break


        # 정보 가져오기
        gun = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/div[3]/section[1]/table[1]/tbody/tr[5]/td[1]").text

        yong = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/div[3]/section[1]/table[1]/tbody/tr[5]/td[2]").text

        saedae = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/div[3]/section[1]/table[1]/tbody/tr[5]/td[3]").text

        parking_num = 0
        for i in range(1,5):
            txt = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/div[3]/div[3]/section/table[1]/tbody/tr/td["+str(i)+"]").text
            txt = re.split("\n", txt)[0]
            txt = re.split("대", txt)[0]
            parking_num = parking_num + int(txt)

        saedae = re.split("세대", saedae)[0].strip() + "세대"

        # if int(re.split("세대", saedae)[0]) == 0:
        #     saedae = browser.find_element_by_xpath("/html/body/div/div/div/div[2]/div[8]/div/div/div/div[3]/section[1]/table[1]/tbody/tr[5]/td[3]").text
        #     saedae = re.split("세대", saedae)[0].strip() + "세대"

        parking_per = parking_num / int(re.split("세대", saedae)[0].strip())
        parking_per = round(parking_per, 2)
        parking = str(parking_num) + "대(세대당" + str(parking_per) + "대)"
        yong = str(re.split('%',yong)[0]).strip() + "%"
        gun = str(re.split('%',gun)[0]).strip() + "%"

        return gun, yong, saedae, parking
    except:
        err_list.append(street_name)
        return 0,0,0,0


for i in range(0, len(street_list)):
    gun, yong, saedae, parking = Search_real(street_list[i])
    print(i)
    if gun == 0:
        ws.cell(row = row_list[i], column= 11, value = "Value_Error")
        ws.cell(row = row_list[i], column= 12, value = "Value_Error")
        ws.cell(row = row_list[i], column= 13, value = "Value_Error")
        ws.cell(row = row_list[i], column= 14, value = "Value_Error")
    else:
        ws.cell(row = row_list[i], column= 11, value = saedae)
        ws.cell(row = row_list[i], column= 12, value = parking)
        ws.cell(row = row_list[i], column= 13, value = yong)
        ws.cell(row = row_list[i], column= 14, value = gun)
    ws.cell(row = row_list[i], column= 15, value = "OK")

import pickle
with open("서울부동산정보광장에러.pickle", "wb") as pickle_data:
    pickle.dump(err_list, pickle_data)


wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/4_도로별_데이터_추가.xlsx")
wb.close()
