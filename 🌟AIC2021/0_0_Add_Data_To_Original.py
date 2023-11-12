from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/6_도로별_데이터_추가.xlsx")
whole_data = {}
ws = wb.active

for i in range(2, ws.max_row + 1):
    data_list = []
    data_list.append(ws.cell(row = i, column = 3).value) # whole_data["street"][0] = mid
    data_list.append(ws.cell(row = i, column = 4).value) # whole_data["street"][1] = high
    data_list.append(ws.cell(row = i, column = 5).value) # whole_data["street"][2] = subway in 500m
    data_list.append(ws.cell(row = i, column = 8).value) # whole_data["street"][3] = 중학교 뺄 년도
    data_list.append(ws.cell(row = i, column = 9).value) # whole_data["street"][4] = 고등학교 뺄 년도
    data_list.append(ws.cell(row = i, column = 11).value) # whole_data["street"][5] = 지하철 뺄 년도
    data_list.append(ws.cell(row = i, column = 12).value) # whole_data["street"][6] = 세대수
    data_list.append(ws.cell(row = i, column = 13).value) # whole_data["street"][7] = 주차공간
    data_list.append(ws.cell(row = i, column = 14).value) # whole_data["street"][8] = 용적률
    data_list.append(ws.cell(row = i, column = 15).value) # whole_data["street"][9] = 건폐율
    whole_data[ws.cell(row = i, column = 1).value] = data_list

import re
err = []
for year in range(2012, 2013):
    print(year)
    wb = load_workbook("/Users/kunwoosmac/Desktop/AIC2021/Data/아파트(전월세)_실거래가_"+str(year)+".xlsx")
    ws = wb.active
    crit = ["중학교",'고등학교','지하철','세대수','주차','용적률','건폐율']
    for i in range(15, 22):
        ws.cell(row = 17, column = i, value = crit[i-15])
    for i in range(18, ws.max_row+1):
        print(i)
        try:
            # 중학교
            try:
                lst = list(re.split('/', str(whole_data[ws.cell(row = i, column = 14).value][4])))
                lst = list(filter(None, lst))
                sub = 0
                for j in range(0, len(lst)):
                    if year > int(lst[j]):
                        sub = sub - 1
                ws.cell(row = i, column = 15, value = int(whole_data[ws.cell(row = i, column = 14).value][0]) + sub)
            except:
                ws.cell(row = i, column = 15, value = int(whole_data[ws.cell(row = i, column = 14).value][0]))

            # 고등학교
            try:
                lst = list(re.split('/', str(whole_data[ws.cell(row = i, column = 14).value][4])))
                lst = list(filter(None, lst))
                sub = 0
                for j in range(0, len(lst)):
                    if year > int(lst[j]):
                        sub = sub - 1
                ws.cell(row = i, column = 16, value = int(whole_data[ws.cell(row = i, column = 14).value][1]) + sub)
            except:
                ws.cell(row = i, column = 16, value = int(whole_data[ws.cell(row = i, column = 14).value][1]))

            # 지하철
            try:
                lst = list(re.split('/', str(whole_data[ws.cell(row = i, column = 14).value][5])))
                lst = list(filter(None, lst))
                sub = 0
                for j in range(0, len(lst)):
                    if year > int(lst[j]):
                        sub = sub - 1
                ws.cell(row = i, column = 17, value = int(whole_data[ws.cell(row = i, column = 14).value][2]) + sub)
            except:
                if whole_data[ws.cell(row = i, column = 14).value][2] != None:
                    ws.cell(row = i, column = 16, value = int(whole_data[ws.cell(row = i, column = 14).value][2]))
                else:
                    ws.cell(row = i, column = 16, value = 0)

            # 세대수, 주차공간, 용적률, 건폐율
            # whole_data["street"][6] = 세대수
            # whole_data["street"][7] = 주차공간
            # whole_data["street"][8] = 용적률
            # whole_data["street"][9] = 건폐율
            for j in range(18, 22):
                ws.cell(row = i, column = j, value = whole_data[ws.cell(row = i, column = 14).value][j - 12])
        except:
            err.append(i)
            ws.cell(row = i, column = 22, value = "도로명 빔")

    wb.save("/Users/kunwoosmac/Desktop/AIC2021/Data/"+str(year)+"_Final.xlsx")
    wb.close()