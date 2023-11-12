from openpyxl import Workbook
from openpyxl import load_workbook
import re
import requests # 웹사이트에서 정보를 받아오기 위한 라이브러리

res = requests.get("https://realty.daum.net/home/apt/map")
res.raise_for_status()
print(res)

wb = Workbook() # 새 워크북을 생성 : 엑셀 파일을 생성함
year = 2012
for year in range(2012, 2022):
    wb = load_workbook("아파트(전월세)_실거래가_"+ str(year) + ".xlsx")
    ws = wb.active
    regex = "\(.*\)|\s-\s.*"
    for x in range(18, ws.max_row + 1):
        get_name = str(ws.cell(row = x, column = 5).value)
        get_name = re.sub(regex,'', get_name)




wb.close()