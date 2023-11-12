from openpyxl import load_workbook

wb = load_workbook("/Users/kunwoosmac/Downloads/도로별_학군추가_에러수정본.xlsx")
ws = wb.active
err_list = []
for i in range(1, ws.max_row + 1):
    if ws.cell(row = i, column = 2).value == "N/A" or \
        ws.cell(row = i, column = 3).value == "N/A" or ws.cell(row = i, column = 4).value == "N/A" or\
            ws.cell(row = i, column = 6).value:
            err_list.append(ws.cell(row = i, column = 1).value)

print(err_list)
import pickle
with open("에러난도로명리스트.pickle", "wb") as pickle_data:
    pickle.dump(err_list, pickle_data)

